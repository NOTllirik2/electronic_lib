import sys, os
import model
import bookInfo
import transact
import profileInfo
import read
import adminButtons
import textRed
import createCopy

import json
from openpyxl import Workbook
from PyQt5 import QtWidgets, QtCore
from datetime import date
from PyQt5.QtGui import QFont, QColor, QBrush
from PyQt5.QtWidgets import QApplication, QTableWidget, QTableWidgetItem, QMainWindow, QLabel, QPushButton, QLineEdit, \
    QApplication, QWidget, QVBoxLayout, QCheckBox, QScrollArea, QComboBox, QTextEdit, QToolBar, QAction, QFileDialog, \
    QHBoxLayout, QMessageBox, QHeaderView
from PyQt5.QtCore import pyqtSignal, QSize, QPropertyAnimation, QRect, Qt
from functools import partial
from model import session, Base, Book, User, Comment, Author, Transaction, book_author_association, user_book_association, book_genre_association

from sqlalchemy import (
    Column, Integer, String, Date, BigInteger, ForeignKey, CheckConstraint,
    Boolean, DECIMAL, UniqueConstraint, create_engine, insert, delete, and_, or_, select, update, func, select, desc)
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from typing import Optional, List, Tuple
from sqlalchemy.exc import SQLAlchemyError


class mainElements(QWidget):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.user = parent.user
        self.setGeometry(0, 0, 1500, 900)

        # Layout для меню
        self.menu_layout = QVBoxLayout(self)
        self.setLayout(self.menu_layout)

        # Элемент QLabel для фона (если он нужен)
        self.background = QLabel(self)
        self.background.setStyleSheet("background-color: rgb(221, 221, 221);")
        self.menu_layout.addWidget(self.background)

        self.label = QLabel("Пользователь авторизовался как " + self.parent.user.role, self)
        self.label.setGeometry(1200, 850, 400, 30)

        self.profileButton = QPushButton("Профиль", self)
        self.profileButton.setGeometry(10, 50, 200, 60)
        self.profileButton.setFont(QFont("Arial", 13))
        self.profileButton.clicked.connect(self.profileClicked)
        self.profileButton.setStyleSheet("background-color: white;")

        self.catalogButton = QPushButton("Каталог", self)
        self.catalogButton.setGeometry(10, 150, 200, 60)
        self.catalogButton.setFont(QFont("Arial", 13))
        self.catalogButton.clicked.connect(self.catalogClicked)
        self.catalogButton.setStyleSheet("background-color: lightBlue;")

        self.biblButton = QPushButton("Библиотека", self)
        self.biblButton.setGeometry(10, 250, 200, 60)
        self.biblButton.setFont(QFont("Arial", 13))
        self.biblButton.clicked.connect(self.biblClicked)
        self.biblButton.setStyleSheet("background-color: white;")

        self.catalogWidget = QWidget(self)
        self.catalogWidget.setGeometry(250, 0, 1500, 900)

        self.profileWidget = profileInfo.profileInfo(self)
        self.profileWidget.setGeometry(250, 0, 1500, 900)
        self.profileWidget.setVisible(False)

        self.biblWidget = QWidget(self)
        self.biblWidget.setGeometry(250, 0, 1500, 900)
        self.biblWidget.setVisible(False)

        self.catalogTable = QTableWidget(self.catalogWidget)
        self.catalogTable.verticalHeader().setVisible(False)
        self.catalogTable.setSelectionMode(QTableWidget.NoSelection)
        self.catalogTable.setEditTriggers(QTableWidget.NoEditTriggers)

        self.calalogLabel = QLabel("Каталог:", self.catalogWidget)
        self.calalogLabel.setGeometry(0, 100, 200, 30)
        self.calalogLabel.setFont(QFont("Arial", 13))

        self.isCatalog = 0
        self.headers = ["Название", "Автор", "Жанр", "Дата выхода", "Цена", "Информация", "Купить"]
        self.catalogTable.setGeometry(0, 140, 1200, 700)
        self.catalogTable.setVerticalHeaderLabels(self.headers)
        self.catalogTable.setStyleSheet("""
                    QTableWidget {
                        border: none;
                        gridline-color: transparent;  /* Убираем вертикальные линии */
                    }
                    QTableWidget::item {
                        border-bottom: 1px solid black;  /* Оставляем только горизонтальные линии */
                    }
                """)

        book_authors = (
            model.session.query(
                model.book_author_association.c.book_id,
                func.group_concat(model.Author.name).label('authors')
            )
            .join(model.Author, model.Author.id == model.book_author_association.c.author_id)
            .group_by(model.book_author_association.c.book_id)
            .subquery()
        )

        book_genres = (
            model.session.query(
                model.book_genre_association.c.book_id,
                func.group_concat(model.Genre.name).label('genres')
            )
            .join(model.Genre, model.Genre.id == model.book_genre_association.c.genre_id)
            .group_by(model.book_genre_association.c.book_id)
            .subquery()
        )

        table = (
            model.session.query(
                model.Book.title,
                book_authors.c.authors,
                book_genres.c.genres,
                model.Book.publication_date,
                model.Book.price
            )
            .outerjoin(book_authors, book_authors.c.book_id == model.Book.id)
            .outerjoin(book_genres, book_genres.c.book_id == model.Book.id)
            .all()
        )

        self.getTable2(self.catalogTable, table)
        self.add_button_column(self.catalogTable, "Информация", self.infoButtonClicked)
        self.add_button_column(self.catalogTable, "Купить", self.buyButtonClicked)

        self.searchLabel = QLabel("Поиск:", self.catalogWidget)
        self.searchLabel.setGeometry(0, 20, 200, 30)
        self.searchLabel.setFont(QFont("Arial", 13))

        self.searchText = QTextEdit(self.catalogWidget)
        self.searchText.setGeometry(0, 50, 200, 30)
        self.searchText.textChanged.connect(self.searchText_changed)

        # ================================================================
        self.biblTable = QTableWidget(self.biblWidget)
        self.biblTable.verticalHeader().setVisible(False)
        self.biblTable.setEditTriggers(QTableWidget.NoEditTriggers)

        self.biblLabel = QLabel("Библиотека:", self.biblWidget)
        self.biblLabel.setGeometry(0, 100, 200, 30)
        self.biblLabel.setFont(QFont("Arial", 13))

        self.headers = ["Название", "Автор", "Жанр", "Дата выхода", "Цена", "Информация", "Купить"]
        self.biblTable.setGeometry(0, 140, 1200, 700)
        self.biblTable.setVerticalHeaderLabels(self.headers)

        self.biblTable.setSelectionMode(QTableWidget.NoSelection)
        self.biblTable.setEditTriggers(QTableWidget.NoEditTriggers)

        self.biblTable.setStyleSheet("""
                        QTableWidget {
                            border: none;
                            gridline-color: transparent;  /* Убираем вертикальные линии */
                        }
                        QTableWidget::item {
                            border-bottom: 1px solid black;  /* Оставляем только горизонтальные линии */
                        }
                    """)

        book_authors = (
            model.session.query(
                model.book_author_association.c.book_id,
                func.group_concat(model.Author.name).label('authors')
            )
            .join(model.Author, model.Author.id == model.book_author_association.c.author_id)
            .group_by(model.book_author_association.c.book_id)
            .subquery()
        )

        book_genres = (
            model.session.query(
                model.book_genre_association.c.book_id,
                func.group_concat(model.Genre.name).label('genres')
            )
            .join(model.Genre, model.Genre.id == model.book_genre_association.c.genre_id)
            .group_by(model.book_genre_association.c.book_id)
            .subquery()
        )

        table = (
            model.session.query(
                model.Book.title,
                book_authors.c.authors,
                book_genres.c.genres,
                model.Book.publication_date,
                model.Book.price
            )
            .outerjoin(book_authors, book_authors.c.book_id == model.Book.id)
            .outerjoin(book_genres, book_genres.c.book_id == model.Book.id)
            .join(model.user_book_association, model.user_book_association.c.book_id == model.Book.id)
            .filter(model.user_book_association.c.user_id == self.user.id)  # Фильтрация по user_id
            .all()
        )

        self.getTable2(self.biblTable, table)
        self.add_button_column(self.biblTable, "Информация", self.infoButtonClicked)
        self.add_button_column(self.biblTable, "Читать", self.readButtonClicked)

        self.searchLabelbibl = QLabel("Поиск:", self.biblWidget)
        self.searchLabelbibl.setGeometry(0, 20, 200, 30)
        self.searchLabelbibl.setFont(QFont("Arial", 13))

        self.searchTextbibl = QTextEdit(self.biblWidget)
        self.searchTextbibl.setGeometry(0, 50, 200, 30)
        self.searchTextbibl.textChanged.connect(self.searchTextBibl_changed)

        book = model.session.query(model.Book).filter(model.Book.id > 0).first()
        self.bookInfo = bookInfo.bookInfo(self, book)
        self.bookInfo.setVisible(False)

        self.transactMenu = transact.buyMenu(self, book)
        self.transactMenu.setVisible(False)

        self.readEl = read.readEl(self, book)
        self.readEl.setVisible(False)

        self.editRow = 0
        self.editCol = 0
        self.editTable = None
        self.startRow = 0
        self.userEdeting = 0
        self.rowSelect = -1
        self.isCatalog = 1

        if self.user.role == "admin":
            self.textRed = textRed.readEl(self, book)
            self.textRed.setVisible(False)

            self.adminTableWidget = QTableWidget(self)
            self.adminTableWidget.setGeometry(250, 50, 1200, 820)
            self.adminTableWidget.setVisible(False)

            self.saveToXlsButton = QPushButton("Сохранить xlsx", self.adminTableWidget)
            self.saveToXlsButton.setGeometry(1075, 750, 120, 60)
            self.saveToXlsButton.clicked.connect(self.saveToXlsButtonClicked)

            self.saveToJsonButton = QPushButton("Сохранить json", self.adminTableWidget)
            self.saveToJsonButton.setGeometry(925, 750, 120, 60)
            self.saveToJsonButton.clicked.connect(self.saveToJsonButtonClicked)

            self.simpleTableButton = QPushButton("Исходные таблицы", self)
            self.simpleTableButton.setGeometry(10, 350, 200, 60)
            self.simpleTableButton.clicked.connect(self.simpleTableButton_clicked)

            self.adminTable = QTableWidget(self.adminTableWidget)
            self.adminTable.setGeometry(0, 120, 1200, 620)
            self.adminTable.itemChanged.connect(self.adminTableItem_changed)
            self.adminTable.cellClicked.connect(self.startEditing)
            self.adminTable.selectionModel().selectionChanged.connect(self.selectionChanged)

            self.addButton = QPushButton("+", self.adminTableWidget)
            self.addButton.setGeometry(10, 750, 120, 60)
            self.addButton.clicked.connect(self.addButton_clicked)

            headers_ = ["Книги", "Авторы", "Транзакции", "Коментарии", "Жанры", "Книги_Авторы", "Книги_Жанры",
                        "Книги_пользователи"]
            x = 0
            h = 0
            for i in range(len(headers_)):
                if headers_[i] == "Книги_Авторы":
                    h += 60
                    x = 0
                self.adminButton = QPushButton(headers_[i], self.adminTableWidget)
                self.adminButton.setGeometry(x, h, 150, 60)
                self.adminButton.clicked.connect(lambda _, i=i: self.selectSimpleTableButtons(headers_[i]))
                x += 150

            self.adminTableForStatisticWidget = QWidget(self)
            self.adminTableForStatisticWidget.setGeometry(250, 50, 1200, 820)
            self.adminTableForStatisticWidget.setVisible(False)

            self.statsTableButton = QPushButton("Таблицы для статистики", self)
            self.statsTableButton.setGeometry(10, 440, 200, 60)
            self.statsTableButton.clicked.connect(self.statsTableButton_clicked)
            # ==========
            self.statsadminTable = QTableWidget(self.adminTableForStatisticWidget)
            self.statsadminTable.setGeometry(0, 70, 1200, 670)

            self.statssaveToXlsButton = QPushButton("Сохранить xlsx", self.adminTableForStatisticWidget)
            self.statssaveToXlsButton.setGeometry(1075, 750, 120, 60)
            self.statssaveToXlsButton.clicked.connect(self.saveToXlsButtonClickedstats)

            self.statssaveToJsonButton = QPushButton("Сохранить json", self.adminTableForStatisticWidget)
            self.statssaveToJsonButton.setGeometry(925, 750, 120, 60)
            self.statssaveToJsonButton.clicked.connect(self.saveToJsonButtonClickedstats)

            headers2_ = ["Популярные книги", "Прибыльные книги", "Прибыльные пользователи", "Активные пользователи",
                        "Популярные авторы"]

            x = 0
            h = 0
            for i in range(len(headers2_)):
                self.statsadminButton = QPushButton(headers2_[i], self.adminTableForStatisticWidget)
                self.statsadminButton.setGeometry(x, h, 200, 70)
                self.statsadminButton.clicked.connect(lambda _, i=i: self.getStatsTable(headers2_[i]))
                x += 200

            self.getCopyButton = QPushButton("Создать резервную копию", self)
            self.getCopyButton.setGeometry(10, 800, 200, 60)
            self.getCopyButton.clicked.connect(self.getCopyButton_clicked)


            self.simpleTableButton.setFont(QFont("Arial", 13))
            self.simpleTableButton.setStyleSheet("background-color: white;")

            self.statsTableButton.setFont(QFont("Arial", 10))
            self.statsTableButton.setStyleSheet("background-color: white;")

    def getCopyButton_clicked(self):
        backup_dir = QFileDialog.getExistingDirectory(self, "Выберите папку для резервной копии", "C:/")
        try:
            path = createCopy.backup_mysql_db("root", "6547", "my_library", backup_dir)
            createCopy.send_file(file_path=path)

            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Резервная копия")
            msg.setText("Резервная копия базы данных успешно создана.")
            msg.exec_()
        except Exception as e:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle("Ошибка")
            msg.setText(f"Ошибка при создании резервной копии: {e}")
            msg.exec_()

    def saveToJsonButtonClickedstats(self):
        self.export_qtable_to_json(self.statsadminTable)

    def saveToXlsButtonClickedstats(self):
        self.export_qtable_to_xlsx(self.statsadminTable)

    def getStatsTable(self, headers_):
        print(headers_)

        if headers_ == "Популярные книги":
            table = (session.query(
                Book.title.label("book_title"),
                func.group_concat(Author.name.distinct()).label("authors"),  # Убираем дубли авторов
                func.count(Transaction.user_id.distinct()).label("purchase_count"),  # Убираем дубли пользователей
                func.sum(Transaction.amount).label("total_revenue")  # Суммируем корректно
            )
                     .join(book_author_association, Book.id == book_author_association.c.book_id)
                     .join(Author, book_author_association.c.author_id == Author.id)
                     .outerjoin(Transaction, Book.id == Transaction.book_id)
                     .group_by(Book.id, Book.title)
                     .order_by(desc("purchase_count"))
                     ).all()

        elif headers_ == "Прибыльные книги":
            table = (session.query(
                Book.title.label("book_title"),
                func.group_concat(Author.name.distinct()).label("authors"),
                func.count(Transaction.user_id.distinct()).label("purchase_count"),
                func.sum(Transaction.amount).label("total_revenue")
            )
                     .join(book_author_association, Book.id == book_author_association.c.book_id)
                     .join(Author, book_author_association.c.author_id == Author.id)
                     .outerjoin(Transaction, Book.id == Transaction.book_id)
                     .group_by(Book.id, Book.title)
                     .order_by(desc("total_revenue"))
                     ).all()

        elif headers_ == "Прибыльные пользователи":
            table = (
                model.session.query(
                    model.User.username.label("user_name"),
                    func.round(func.sum(model.Transaction.amount), 2).label("total_spent")
                )
                .join(model.Transaction, model.User.id == model.Transaction.user_id)
                .group_by(model.User.id, model.User.username)
                .order_by(desc("total_spent"))
                .all()
            )
        elif headers_ == "Активные пользователи":
            table = (
                model.session.query(
                    model.User.username.label("user_name"),
                    func.count(model.user_book_association.c.book_id).label("books_count")
                )
                .join(model.user_book_association, model.User.id == model.user_book_association.c.user_id)
                .group_by(model.User.id, model.User.username)
                .order_by(desc("books_count"))
                .all()
            )
        elif headers_ == "Популярные авторы":
            table = (
                model.session.query(
                    model.Author.name.label("author_name"),
                    func.count(model.user_book_association.c.book_id).label("books_purchased")
                )
                .join(model.book_author_association, model.Author.id == model.book_author_association.c.author_id)
                .join(model.user_book_association, model.book_author_association.c.book_id == model.user_book_association.c.book_id)
                .group_by(model.Author.id, model.Author.name)
                .order_by(desc("books_purchased"))
                .all()
            )

        self.getTable3(self.statsadminTable, table)

    def statsTableButton_clicked(self):
        self.setButtonBlue(3)
        self.setWidget(3)

    def saveToXlsButtonClicked(self):
        self.export_qtable_to_xlsx(self.adminTable)

    def saveToJsonButtonClicked(self):
        self.export_qtable_to_json(self.adminTable)

    def export_qtable_to_json(self, table_widget):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Сохранить файл",
            "",
            "JSON Files (*.json)",
            options=options
        )

        if not file_path:
            return

        table_data = []

        column_count = table_widget.columnCount()
        headers = [
            table_widget.horizontalHeaderItem(col).text() if table_widget.horizontalHeaderItem(
                col) else f"Column {col + 1}"
            for col in range(column_count)
        ]

        row_count = table_widget.rowCount()
        for row in range(row_count):
            row_data = {}
            for col in range(column_count):
                header = headers[col]
                item = table_widget.item(row, col)
                row_data[header] = item.text() if item else None
            table_data.append(row_data)

        try:
            with open(file_path, 'w', encoding='utf-8') as json_file:
                json.dump(table_data, json_file, ensure_ascii=False, indent=4)
            print(f"Таблица успешно сохранена в файл: {file_path}")
        except Exception as e:
            print(f"Ошибка при сохранении файла: {e}")

    def export_qtable_to_xlsx(self, table_widget):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Сохранить файл",
            "",
            "Excel Files (*.xlsx)",
            options=options
        )

        if not file_path:
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Exported Data"

        column_count = table_widget.columnCount()
        row_count = table_widget.rowCount()

        for col in range(column_count):
            header_item = table_widget.horizontalHeaderItem(col)
            if header_item:
                sheet.cell(row=1, column=col + 1, value=header_item.text())

        for row in range(row_count):
            for col in range(column_count):
                item = table_widget.item(row, col)
                if item:
                    sheet.cell(row=row + 2, column=col + 1, value=item.text())

        try:
            workbook.save(file_path)
            print(f"Таблица успешно сохранена в файл: {file_path}")
        except Exception as e:
            print(f"Ошибка при сохранении файла: {e}")

    def tryToAddData(self):
        if self.editTable is None:
            return

        name = getattr(self.editTable, 'name', None) or getattr(self.editTable, '__tablename__', None)

        if str(name) == "Author.name":
            name = "authors"

        if str(name) == "Genre.name":
            name = "genres"

        for i in range(self.startRow, self.adminTable.rowCount()):
            r = self.adminTable.columnCount()

            if name == "books":
                r -= 1

            fields = []
            data = []
            for j in range(r):
                item = self.adminTable.item(i, j)
                if not (item is None) and not (item.text() == ""):
                    fields.append(self.adminTable.horizontalHeaderItem(j).text())
                    data.append(item.text())

            try:
                model.addTable(name, fields, [data])
            except ValueError as ve:
                print(f"Ошибка: {ve}")
            except SQLAlchemyError as se:
                print(f"Ошибка работы с базой данных: {se}")
            except Exception as e:
                print(f"Непредвиденная ошибка: {e}")

    def addButton_clicked(self):
        self.addRow(self.adminTable)

    def startEditing(self, row, col):
        self.editRow = row
        self.editCol = col
        self.userEdeting = 1

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Delete:
            self.delete_selected_cells()
        else:
            super().keyPressEvent(event)

    def delete_selected_cells(self):
        if self.rowSelect == -1:
            return

        print("Удаление строки " + str(self.rowSelect))
        name = getattr(self.editTable, 'name', None) or getattr(self.editTable, '__tablename__', None)

        if str(name) == "Author.name":
            name = "authors"

        if str(name) == "Genre.name":
            name = "genres"

        r = self.adminTable.columnCount()

        if name == "books":
            r -= 1

        fields = []
        data = []
        for j in range(r):
            print(j)
            item = self.adminTable.item(self.rowSelect, j)
            if not (item is None) \
                    and not (item.text() == "") \
                    and not (item.text() == "None") \
                    and not (self.adminTable.horizontalHeaderItem(j).text() == "price") \
                    and not (self.adminTable.horizontalHeaderItem(j).text() == "amount"):
                fields.append(self.adminTable.horizontalHeaderItem(j).text())
                data.append(item.text())

        print(name)
        print(fields)
        print(data)

        if len(fields) != 0 and len(data) != 0:
            try:
                model.dellTable(name, fields, data)
            except ValueError as ve:
                print(f"Ошибка: {ve}")
            except SQLAlchemyError as se:
                print(f"Ошибка работы с базой данных: {se}")
            except Exception as e:
                print(f"Непредвиденная ошибка: {e}")

        self.selectSimpleTableButtons(self.tableName)

    def selectionChanged(self, selected, deselected):
        selected_indexes = self.adminTable.selectedIndexes()

        if self.isSingleRowSelected(selected_indexes):
            self.rowSelect = selected_indexes[0].row()
        else:
            self.rowSelect = -1

    def isSingleRowSelected(self, selected_indexes):
        rows = [index.row() for index in selected_indexes]
        return len(set(rows)) == 1 and len(selected_indexes) == self.adminTable.columnCount()

    def adminTableItem_changed(self):
        if self.editTable is None:
            return

        if self.userEdeting == 0:
            return

        self.userEdeting = 0

        print(self.editRow, self.startRow)
        if self.editRow >= self.startRow:
            return


        name = getattr(self.editTable, 'name', None) or getattr(self.editTable, '__tablename__', None)

        if str(name) == "Author.name":
            name = "authors"

        if str(name) == "Genre.name":
            name = "genres"

        r = self.adminTable.columnCount()

        if name == "books":
            r -= 1

        fields = []
        data = []
        for j in range(r):
            print(j)
            if j != self.editCol:
                item = self.adminTable.item(self.editRow, j)
                if not (item is None) \
                        and not (item.text() == "") \
                        and not (item.text() == "None") \
                        and not (self.adminTable.horizontalHeaderItem(j).text() == "price") \
                        and not (self.adminTable.horizontalHeaderItem(j).text() == "amount"):
                    fields.append(self.adminTable.horizontalHeaderItem(j).text())
                    data.append(item.text())

        item = self.adminTable.item(self.editRow, self.editCol)
        if item is None:
            return
        dataToAppend = item.text()
        item = self.adminTable.horizontalHeaderItem(self.editCol)

        if item is None:
            return

        fieldToAppend = item.text()

        print(name)
        print(fields)
        print(data)
        print(fieldToAppend)
        print(dataToAppend)

        if len(fields) != 0 and len(data) != 0:
            try:
                model.updateTable(name, fieldToAppend, dataToAppend, fields, data)
            except ValueError as ve:
                print(f"Ошибка: {ve}")
            except SQLAlchemyError as se:
                print(f"Ошибка работы с базой данных: {se}")
            except Exception as e:
                print(f"Непредвиденная ошибка: {e}")

    def addRow(self, table):
        rowCount = table.rowCount()
        table.insertRow(rowCount)
        for i in range(1, table.columnCount()):
            table.setItem(rowCount, i, QTableWidgetItem(""))

    def selectSimpleTableButtons(self, tableName):
        self.tryToAddData()
        self.tableName = tableName

        table = (
            model.session.query(model.Book.id, model.Book.title, model.Book.publication_date, model.Book.description,
                                model.Book.price).all())

        if tableName == "Книги":
            self.editTable = model.Book
            table = (
                model.session.query(model.Book.id, model.Book.title, model.Book.publication_date,
                                    model.Book.description,
                                    model.Book.price).all())
        elif tableName == "Авторы":
            self.editTable = model.Author
            table = (
                model.session.query(model.Author.id, model.Author.name, model.Author.biography).all())
        elif tableName == "Транзакции":
            self.editTable = model.Transaction
            table = (
                model.session.query(model.Transaction.id, model.Transaction.user_id, model.Transaction.book_id,
                                    model.Transaction.amount, model.Transaction.transaction_type,
                                    model.Transaction.card_num).all())
        elif tableName == "Коментарии":
            self.editTable = model.Comment
            table = (model.session.query(model.Comment.id, model.Comment.user_id, model.Comment.book_id,
                                         model.Comment.content).all())
        elif tableName == "Жанры":
            self.editTable = model.Genre
            table = (model.session.query(model.Genre.id, model.Genre.name).all())

        elif tableName == "Книги_Авторы":
            self.editTable = model.book_author_association
            table = model.session.query(model.book_author_association.c.book_id,
                                        model.book_author_association.c.author_id).all()
        elif tableName == "Книги_Жанры":
            self.editTable = model.book_genre_association
            table = model.session.query(model.book_genre_association.c.book_id,
                                        model.book_genre_association.c.genre_id).all()
        elif tableName == "Книги_пользователи":
            self.editTable = model.user_book_association
            table = model.session.query(model.user_book_association.c.user_id,
                                        model.user_book_association.c.book_id).all()
        self.adminTable.clear()
        self.getTable3(self.adminTable, table)

        if tableName == "Книги":
            he = ['id', 'title', 'publication_date', 'description', 'price', 'content']
            self.add_button_column(self.adminTable, "Текст", self.textLoad)
            self.adminTable.setHorizontalHeaderLabels(he)

        for i in range(self.adminTable.rowCount()):
            self.adminTable.setRowHeight(i, 50)

        self.startRow = self.adminTable.rowCount()

    def textLoad(self, row):
        print(row)
        id = int(self.adminTable.item(row, 0).text())
        book = model.session.query(model.Book).filter(model.Book.id == id).first()

        self.textRed.book = book
        self.textRed.loadInfo()
        self.textRed.setVisible(True)

    def simpleTableButton_clicked(self):
        self.setButtonBlue(4)
        self.setWidget(4)

    def setButtonBlue(self, index):
        buttons = [self.profileButton, self.catalogButton, self.biblButton]
        if self.user.role == "admin":
            buttons.append(self.statsTableButton)
            buttons.append(self.simpleTableButton)

        for button in buttons:
            button.setStyleSheet("background-color: white;")

        if index != -1:
            buttons[index].setStyleSheet("background-color: lightBlue;")

    def setWidget(self, index):
        widgets = [self.profileWidget, self.catalogWidget, self.biblWidget]
        if self.user.role == "admin":
            widgets.append(self.adminTableForStatisticWidget)
            widgets.append(self.adminTableWidget)

        for widget in widgets:
            widget.setVisible(False)

        widgets[index].setVisible(True)

    def readButtonClicked(self, row):
        book = self.getBookFromRow(row)
        self.readEl.book = book
        self.readEl.loadInfo()
        self.readEl.setVisible(True)

    def profileClicked(self):
        self.setButtonBlue(0)
        self.setWidget(0)

    def catalogClicked(self):
        self.setButtonBlue(1)
        self.setWidget(1)
        self.isCatalog = 1

    def biblClicked(self):
        self.setButtonBlue(2)
        self.setWidget(2)
        self.isCatalog = 0

        self.biblTable.hideColumn(4)
        self.biblTable.setHorizontalHeaderLabels(["Название", "Автор", "Жанр", "Дата выхода", "", "Информация", "Читать"])

    def getBookFromRow(self, row):
        if self.isCatalog:
            name = self.catalogTable.item(row, 0).text()
        else:
            name = self.biblTable.item(row, 0).text()

        return model.session.query(model.Book).filter(model.Book.title == name).first()

    def buyButtonClicked(self, row):
        book = self.getBookFromRow(row)
        print(book)
        self.transactMenu.book = book
        if self.transactMenu.loadInfo():
            self.transactMenu.setVisible(True)

    def infoButtonClicked(self, row):
        book = self.getBookFromRow(row)
        print(book)
        self.bookInfo.book = book
        self.bookInfo.loadInfo()
        self.bookInfo.setVisible(True)

    def searchTextBibl_changed(self):
        text = self.searchTextbibl.toPlainText()
        print(text)
        self.search_in_table(self.biblTable, text)

    def searchText_changed(self):
        text = self.searchText.toPlainText()
        print(text)
        self.search_in_table(self.catalogTable, text)

    def search_in_table(self, table_widget, search_text):
        row_count = table_widget.rowCount()
        column_count = table_widget.columnCount()

        for row_idx in range(row_count):
            row_contains_text = False
            for col_idx in range(2):
                item = table_widget.item(row_idx, col_idx)
                if item and search_text.lower() in item.text().lower():
                    row_contains_text = True
                    break

            table_widget.setRowHidden(row_idx, not row_contains_text)

    def add_button_column(self, table_widget, button_text="Click", button_action=None):
        column_index = table_widget.columnCount()

        table_widget.setColumnCount(column_index + 1)

        for row_idx in range(table_widget.rowCount()):
            button = QPushButton(button_text)

            if button_action:
                button.clicked.connect(lambda _, row=row_idx: button_action(row))

            button_widget = QWidget()
            button_layout = QHBoxLayout(button_widget)
            button_layout.addWidget(button)
            button_layout.setAlignment(Qt.AlignCenter)

            table_widget.setCellWidget(row_idx, column_index, button_widget)
            for row_idx in range(table_widget.rowCount()):
                table_widget.setRowHeight(row_idx, 100)
        table_widget.setHorizontalHeaderLabels(self.headers)

    def getTable3(self, table_widget, query_result):
        if not query_result:
            return

        table_widget.setRowCount(len(query_result))
        table_widget.setColumnCount(len(query_result[0]))

        column_headers = [column for column in query_result[0]._fields]

        for row_idx, row in enumerate(query_result):
            for col_idx, col_value in enumerate(row):
                if col_idx == 1 or col_idx == 2:
                    col_value = str(col_value).replace(",", "\n")

                item = QTableWidgetItem(str(col_value))
                item.setTextAlignment(Qt.AlignCenter)

                table_widget.setItem(row_idx, col_idx, item)

        table_widget.setHorizontalHeaderLabels(column_headers)

        header = table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)

    def getTable2(self, table_widget, query_result):
        if not query_result:
            return

        table_widget.setRowCount(len(query_result))
        table_widget.setColumnCount(len(query_result[0]))

        for row_idx, row in enumerate(query_result):
            for col_idx, col_value in enumerate(row):
                if col_idx == 1 or col_idx == 2:
                    col_value = str(col_value).replace(",", "\n")
                    item = QTableWidgetItem(f"{col_value}")
                    item.setTextAlignment(Qt.AlignCenter)
                    table_widget.setItem(row_idx, col_idx, item)
                else:
                    item = QTableWidgetItem(f"{col_value}")
                    item.setTextAlignment(Qt.AlignCenter)
                    table_widget.setItem(row_idx, col_idx, item)

        header = table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        table_widget.setHorizontalHeaderLabels(self.headers)
